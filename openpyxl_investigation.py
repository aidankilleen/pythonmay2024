import openpyxl

filename = "salesrecords.xlsx"
wb = openpyxl.load_workbook(filename)

print (wb.sheetnames)

ws = wb.create_sheet("Aidans Sheet")

ws.cell(1, 1).value = "Aidans Data Sheet"

for i in range(1, 10):
    ws.cell(i+1, 1).value = i

ws.cell(11,1).value = f"=SUM(A2:A10)"

wb.save(filename)

wb.close()
